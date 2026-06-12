import os
import time
import base64
import urllib.parse
from urllib.parse import urljoin
import requests
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from playwright.sync_api import sync_playwright
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from concurrent.futures import ThreadPoolExecutor
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import threading

BASE_URL = "https://in.ixl.com"
TARGET_URL = "https://www.ixl.com/math/grade-3"
LOGIN_URL = "https://in.ixl.com/signin"
EMAIL = "parkerhouston411@kacad"
PASSWORD = "81party"
QUESTIONS_PER_SKILL = 3
EXCEL_FILENAME = "ixl_grade3_questions.xlsx"
IMAGE_DIR = "ixl_diagrams"
DRIVE_FOLDER_ID = "1ffreALNKiFdOO2dT6Qmunp-vXvO-u3Nm"
SCOPES = ['https://www.googleapis.com/auth/drive']
CLIENT_SECRET_FILE = "client_secret.json"
TOKEN_FILE = "token.json"

# Mode 2: set this to the skill URL you want to resume from
START_URL = "https://www.ixl.com/math/grade-1/giving-to-charity"
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

JUNK_LABELS = {"scratchpad", "eraser", "highlighter - blue",
               "pencil - black", "pencil", "highlighter"}

ICON_MAX = 32

DIAGRAM_SIGNALS = [
    ".multiplication-model-container",
    ".guide-counting-qm",
    ".open-number-line",
    ".dc-fraction-strip-model",
    ".pie-chart",
    ".vector-image-wrapper",
    ".shape",
    "canvas",
    '[role="figure"]',
    "table.old-table",
    "table.qTabularGrid",
    ".qPVTable",
    "svg:has(g.grid-region)",
    "div.table:has([data-testid='area-model-cell'])",
    ".gc-cut-shapes",
    ".fractionTopBlockDiv",
    ".horizontal-scroll-element-wrapper",
    ".horizontal-scroll-hoc-wrapper",
    "table:has(img[src*='~media'])",
    ".guide-counting-clickable-image-container",
    ".standalone-cube-train-wrapper",
    ".hundredTable",
    # New signals
    ".train-and-item-group",
    ".train-and-element-group",
    ".measurementRegion",
    ".calendar-container",
    ".diagramLabelContainer",
    ".QMMeasurable",
    "[class*='tenFrames']",
    "[class*='series-of-components']",
    "[class*='qTable']",
    "[class*='pvmContainer']", 
    "[class*='clockContainer']",
    "[class*='currencyCoinDiv']",
    "[class*='horizontal-scroll']",
    "[class*='dragAndDropContainer']",
    "[class*='story-book']",
    "[class*='static-cube-train']",
    "[class*='SelectableTime']",
    ".simple-item-table"
]

Q_SCOPE_PARTS = [
    ".question-and-submission-view .secContent",
]

# image dimensions in Excel
EXCEL_IMG_MAX_W = 200
EXCEL_IMG_MAX_H = 150
EXCEL_ROW_HEIGHT_PER_IMG = 115

_drive_service = None
_drive_executor = ThreadPoolExecutor(max_workers=5)


def _init_drive_service():
    global _drive_service
    creds = None
    if os.path.exists(TOKEN_FILE):
        creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRET_FILE, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_FILE, 'w') as token:
            token.write(creds.to_json())
    _drive_service = build('drive', 'v3', credentials=creds)


def _create_drive_folder(name, parent_id):
    try:
        metadata = {
            'name': name,
            'mimeType': 'application/vnd.google-apps.folder',
            'parents': [parent_id]
        }
        folder = _drive_service.files().create(body=metadata, fields='id').execute()
        folder_id = folder['id']
        return folder_id
    except Exception as e:
        print(f"     [!] Drive folder creation failed ({name}): {e}")
        return None


_thread_local = threading.local()

def _get_thread_drive_service():
    if not hasattr(_thread_local, 'service'):
        creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)
        _thread_local.service = build('drive', 'v3', credentials=creds)
    return _thread_local.service

def _upload_file_to_drive_sync(file_path, folder_id):
    try:
        service = _get_thread_drive_service()
        metadata = {'name': os.path.basename(file_path), 'parents': [folder_id]}
        media = MediaFileUpload(file_path, mimetype='image/png')
        service.files().create(body=metadata, media_body=media, fields='id').execute()
    except Exception as e:
        print(f"     [!] Drive upload failed ({file_path}): {e}")

def _upload_file_to_drive(file_path, folder_id):
    _drive_executor.submit(_upload_file_to_drive_sync, file_path, folder_id)


def _drive_subfolder_url(folder_id):
    return f"https://drive.google.com/drive/folders/{folder_id}"


def setup_dir():
    os.makedirs(IMAGE_DIR, exist_ok=True)

# Create excel if doesnt exist
def init_excel():
    if os.path.exists(EXCEL_FILENAME):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Grade 3 Maths"
    headers = ["#", "Category", "Skill Name", "Question No",
               "Question Text", "Question Diagram", "Question Options",
               "Option Diagrams", "Correct Answer", "Answer Diagram"]
    header_font = Font(name="Calibri", bold=True, size=11)

    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    widths = {"A": 6, "B": 30, "C": 45, "D": 12,
              "E": 70, "F": 50, "G": 35, "H": 50, "I": 35, "J": 55}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:J1"
    wb.save(EXCEL_FILENAME)


def append_to_excel(row_data, q_diagram_url, opt_diagram_url, ans_diagram_url=None):
    try:
        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        current_row = ws.max_row + 1
        cell_font = Font(name="Calibri", size=11)

        def _clean_str(s):
            if s is None: return ""
            return ILLEGAL_CHARACTERS_RE.sub('', str(s))

        text_columns = {
            1: (_clean_str(row_data[0]), None),
            2: (_clean_str(row_data[1]), None),
            3: (_clean_str(row_data[2]), None),
            4: (_clean_str(row_data[3]), None),
            5: (_clean_str(row_data[4]), None),
            6: (_clean_str(q_diagram_url), _clean_str(q_diagram_url) if q_diagram_url else None),
            7: (_clean_str(row_data[5]), None),
            8: (_clean_str(opt_diagram_url), _clean_str(opt_diagram_url) if opt_diagram_url else None),
            # Col I: plain correct-answer text (never hyperlinked)
            9: (_clean_str(row_data[6]), None),
            # Col J: Drive URL for answer diagram, hyperlinked (empty if no diagram)
            10: (_clean_str(ans_diagram_url) if ans_diagram_url else "",
                 _clean_str(ans_diagram_url) if ans_diagram_url else None),
        }

        for col_idx, (value, h_link) in text_columns.items():
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            if h_link and h_link.startswith("http"):
                cell.hyperlink = h_link
                cell.font = Font(name="Calibri", size=11, color="0563C1", underline="single")
            else:
                cell.font = cell_font
            cell.border = BORDER
            if col_idx in [1, 4]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        if not ws.row_dimensions[current_row].height:
            ws.row_dimensions[current_row].height = 40

        wb.save(EXCEL_FILENAME)
        print(f"       [Excel] Successfully saved Q{row_data[3]} to {EXCEL_FILENAME}")

    except PermissionError:
        print(f"\n[!] Close {EXCEL_FILENAME} in Excel immediately.")
        input("Press ENTER here once closed to resume saving data...")
        append_to_excel(row_data, q_diagram_url, opt_diagram_url, ans_diagram_url)
    except Exception as e:
        print(f"       [!] Failed to append to Excel: {e}")

# jscode to extract text from the question view
_MATH_WALKER_JS = """
    const _DIAGRAM_CLASSES = [
        'dc-fraction-strip-model', 'open-number-line', 'graphingBaseContainer',
        'pie-chart', 'multiplication-model-container', 'guide-counting-qm',
        'vector-image-wrapper', 'parking-lot', 'old-table', 'binsContainer', 'qTabularGrid', 'table',
        'gc-cut-shapes', 'shape', 'fractionTopBlockDiv', 'SelectableTile', 'TileMultipleChoices',
        'ddItemBankDropSlot', 'answer-box', 'react-gc-number-button-grid', 'buttonShape',
        'canvas-container-div', 'tenFrames-single', 'static-cube-train',
        'simple-item-table', 'expression-tile-bank', 'expression-tile-parking-space', 'expression-tile',
        'train-and-item-group', 'train-and-element-group', 'measurementRegion',
        'calendar-container', 'diagramLabelContainer', 'QMMeasurable'
    ];
    const isDiagramTableOrElement = (el) => {
        if (!el.classList) return false;
        if (_DIAGRAM_CLASSES.some(c => el.classList.contains(c))) return true;
        if (Array.from(el.classList).some(c => c.includes('series-of-components') || c.includes('story-book') || c.includes('qTable') || c.includes('pvmContainer') || c.includes('clockContainer') || c.includes('currencyCoinDiv') || c.includes('horizontal-scroll') || c.includes('dragAndDropContainer') || c.includes('SelectableTime') || c.includes('static-cube-train'))) return true;
        if (el.classList.contains('hundredTable') || el.classList.contains('flowerTable')) return true;
        if (el.tagName.toLowerCase() === 'table' && el.querySelector('img[src*="~media"]')) return true;
        if (el.tagName.toLowerCase() === 'canvas') return true;
        if (el.getAttribute('role') === 'figure') return true;
        return false;
    };
    let out = '';
    const walk = (node) => {
        for (const child of node.childNodes) {
            if (child.nodeType === Node.TEXT_NODE) {
                out += child.textContent;
            } else if (child.nodeType === Node.ELEMENT_NODE) {
                if (child.getAttribute('aria-hidden') === 'true') continue;
                const tag = child.tagName.toLowerCase();
                if (tag === 'button') continue;
                // --- Math-expression checks FIRST (before diagram-class skip) ---
                if (tag === 'input' && child.classList.contains('fillIn')) {
                    out += '_';
                } else if (tag === 'div' && child.classList && child.classList.contains('drop-slot')) {
                    out += '_';
                } else if (tag === 'table' && child.hasAttribute('audioalt')) {
                    // Old-style fraction table on option tiles: <table audioalt="2/3">
                    out += child.getAttribute('audioalt');
                } else if (child.classList && child.classList.contains('old-fraction-in-text')) {
                    // Old-style fraction in question text: span.old-fraction-in-text > table
                    const tbl = child.querySelector('table');
                    if (tbl && tbl.hasAttribute('audioalt')) {
                        out += tbl.getAttribute('audioalt');
                    } else if (tbl) {
                        const rows = tbl.querySelectorAll('tr');
                        const num = rows.length > 0 ? rows[0].textContent.trim() : '?';
                        const den = rows.length > 1 ? rows[1].textContent.trim() : '?';
                        out += num + '/' + den;
                    }
                } else if (child.classList && child.classList.contains('vFrac')) {
                    const numEl = child.querySelector('.numerator');
                    const denEl = child.querySelector('.denominator');
                    const numText = (numEl && numEl.querySelector('input.fillIn')) ? '_'
                                  : (numEl ? numEl.textContent.trim() : '?');
                    const denText = (denEl && denEl.querySelector('input.fillIn')) ? '_'
                                  : (denEl ? denEl.textContent.trim() : '?');
                    out += numText + '/' + denText;
                } else if (child.classList && child.classList.contains('vertArith')) {
                    const rows = [...child.querySelectorAll('.vertArithRow')];
                    const operands = [];
                    let operator = '';
                    let answerBlanks = 0;
                    for (const row of rows) {
                        if (row.getAttribute('role') === 'group') {
                            answerBlanks = row.querySelectorAll('input.fillIn').length;
                            continue;
                        }
                        const opCell = row.querySelector('.vertArithCell.operator');
                        if (opCell) {
                            let sym = opCell.textContent.trim();
                            sym = sym === '\\u2013' ? '-' : sym === '\\u00f7' ? '/' : (sym === '\\u00d7' || sym === '×') ? 'x' : sym;
                            if (!operator) operator = sym;
                        } else {
                            const xSym = row.querySelector('.xSymbol');
                            if (xSym) {
                                operator = 'x';
                            }
                        }
                        let numStr = '';
                        for (const cell of row.querySelectorAll('.vertArithCell')) {
                            if (cell.classList.contains('operator')) continue;
                            const txt = cell.querySelector('.txt');
                            if (txt) { numStr += txt.textContent.trim(); }
                            else if (cell.classList.contains('rtlCell')) {
                                const d = cell.textContent.trim();
                                if (d) numStr += d;
                            } else {
                                const expNums = cell.querySelectorAll('.expression.number');
                                for (const expNum of expNums) {
                                    numStr += expNum.textContent.trim();
                                }
                            }
                        }
                        if (numStr) operands.push(numStr);
                    }
                    if (operands.length > 0) {
                        const op = operator || '+';

                        let eq = operands.join(' ' + op + ' ');
                        if (answerBlanks > 0) {
                            eq += ' = ' + Array(answerBlanks).fill('_').join(' ');
                        }
                        if (op === 'x' || op === '\\u00d7' || op === '×') {
                            out += 'Multiply: ' + eq;
                        } else {
                            out += eq;
                        }
                    }
                } else if (child.classList && child.classList.contains('old-vertiArith')) {
                    const rows = [...child.querySelectorAll('table tr')];
                    const entries = [];
                    let operator = '';
                    for (const row of rows) {
                        const cells = [...row.querySelectorAll('td')];
                        let numStr = '';
                        for (let i = 0; i < cells.length; i++) {
                            if (i === 0) {
                                const t = cells[i].textContent.trim();
                                if (t && t !== ' ') {
                                    operator = t === '\\u2013' ? '-' : t === '\\u00f7' ? '/' : t;
                                }
                                continue;
                            }
                            const fi = cells[i].querySelector('input.fillIn');
                            if (fi) { numStr += '_'; }
                            else {
                                const inner = cells[i].querySelector('div') || cells[i];
                                const t = inner.textContent.trim();
                                if (t && t !== ' ') numStr += t;
                            }
                        }
                        if (numStr) entries.push(numStr);
                    }
                    if (entries.length >= 2) {
                        const op = operator || '+';
                        out += entries[0] + ' ' + op + ' ' + entries[1];
                        if (entries.length >= 3) out += ' = ' + entries[2];
                    }
                // --- NOW skip diagram containers (after all math checks) ---
                } else if (isDiagramTableOrElement(child)) {
                    continue;
                } else {
                    walk(child);
                }
            }
        }
    };
"""

"""Only activates for fill-in questions; formats fractions as num/den."""
def _reconstruct_with_blanks(page, root_selector):

    js = f"""
    (sel) => {{
        const root = document.querySelector(sel);
        if (!root) return null;
        if (!root.querySelector('input.fillIn')) return null;
        {_MATH_WALKER_JS}
        walk(root);
        return out;
    }}
    """
    try:
        return page.evaluate(js, root_selector)
    except Exception:
        return None

    """Extract text from selector with fractions as num/den; works for all question types."""
def _extract_math_text(page, root_selector):

    js = f"""
    (sel) => {{
        const root = document.querySelector(sel);
        if (!root) return null;
        {_MATH_WALKER_JS}
        walk(root);
        return out.trim() || null;
    }}
    """
    try:
        return page.evaluate(js, root_selector)
    except Exception:
        return None


def _normalize_math_text(text):
    """Normalize Unicode math symbols to plain-text equivalents."""
    if not text:
        return text
    # Replace Unicode multiplication signs with 'x'
    text = text.replace('×', ' x ')   # ×
    text = text.replace('·', ' x ')   # ·  (middle dot often used for mult)
    text = text.replace('⋅', ' x ')   # ⋅
    # Replace Unicode minus/dash with '-'
    text = text.replace('–', '-')      # –  en-dash
    text = text.replace('−', '-')      # −  minus sign
    # Replace Unicode division sign with '/'
    text = text.replace('÷', '/')      # ÷
    # Collapse multiple spaces
    text = ' '.join(text.split())
    return text


def _get_active_question(page):
    _q_candidates = []
    _seen_q_coords = set()
    for _scope in (".question-and-submission-view", ".ixl-practice-crate"):
        try:
            for _el in page.locator(_scope).all():
                _bb = _safe_bbox(_el)
                if _bb is None or _bb["width"] < 1 or _bb["height"] < 1 or _bb["y"] < 0:
                    continue
                _coord = (round(_bb["x"]), round(_bb["y"]), round(_bb["width"]), round(_bb["height"]))
                if _coord not in _seen_q_coords:
                    _seen_q_coords.add(_coord)
                    _q_candidates.append((_bb["y"], _el))
        except Exception:
            pass

    if _q_candidates:
        _q_candidates.sort(key=lambda t: t[0])
        return _q_candidates[0][1]
    return None


def extract_question_text(page, root_locator=None):
    if root_locator is not None:
        active_q = root_locator
    else:
        active_q = _get_active_question(page)
        if not active_q:
            active_q = page.locator(".ixl-practice-crate").first
            if active_q.count() == 0:
                active_q = page.locator(".question-and-submission-view").first

    if active_q and active_q.count() > 0:
        try:
            js = f"el => {{ {_MATH_WALKER_JS} walk(el); return out.trim(); }}"
            rebuilt = active_q.evaluate(js)
            if rebuilt:
                text = " ".join(rebuilt.replace("\n", " ").split())
                return _normalize_math_text(text)
        except Exception as e:
            print(f"     [!] active_q evaluate failed: {e}")

    # Fallback to old-style logic if active_q evaluation fails or is empty
    if root_locator is not None:
        scopes = [root_locator]
    else:
        scopes = [
            page.locator(".question-and-submission-view .math.section").first,
            page.locator(".question-and-submission-view .ixl-practice-crate").first,
            page.locator(".ixl-practice-crate").first
        ]

    for sel_loc in scopes:
        if sel_loc.count() > 0:
            try:
                js_reconstruct = f"el => {{ if (!el.querySelector('input.fillIn')) return null; {_MATH_WALKER_JS} walk(el); return out; }}"
                rebuilt = sel_loc.evaluate(js_reconstruct)
                if rebuilt is not None:
                    text = " ".join(rebuilt.replace("\n", " ").split())
                    return _normalize_math_text(text)
            except Exception:
                pass

    question_text = ""
    scope = root_locator if root_locator is not None else page
    hdr     = scope.locator(".secHdr").first
    content = scope.locator(".secContent").first
    crate   = root_locator if root_locator is not None else scope.locator(".ixl-practice-crate").first

    try:
        parts = []
        js_extract = f"el => {{ {_MATH_WALKER_JS} walk(el); return out.trim() || null; }}"
        if hdr.count() > 0 and hdr.is_visible():
            try:
                hdr_text = hdr.evaluate(js_extract)
            except Exception:
                hdr_text = None
            parts.append(hdr_text if hdr_text else hdr.inner_text())
        if content.count() > 0 and content.is_visible():
            try:
                content_text = content.evaluate(js_extract)
            except Exception:
                content_text = None
            if content_text and content_text.strip():
                parts.append(content_text)
        if parts:
            question_text = "\n".join(
                " ".join(p.replace("\n", " ").split()) for p in parts if p.strip()
            )
        elif crate.count() > 0 and crate.is_visible():
            try:
                crate_text = crate.evaluate(js_extract)
            except Exception:
                crate_text = None
            question_text = " ".join(
                (crate_text or crate.inner_text()).split("Submit")[0].replace("\n", " ").split()
            )
    except Exception as e:
        print(f"     [!] text read failed: {e}")

    return _normalize_math_text(question_text)


def extract_options(page, root_locator=None):
    options = []

    if root_locator is not None:
        tiles = root_locator.locator(".SelectableTile").all()
    else:
        tiles = page.locator(
            ".question-and-submission-view .SelectableTile, "
            ".ixl-practice-crate .SelectableTile"
        ).all()

    _walker_js = f"el => {{ {_MATH_WALKER_JS} walk(el); return out.trim(); }}"
    _exp_js = """el => {
        if (el.classList.contains('expression-tile-bank') || el.querySelector('.expression-tile-bank')) {
            const spans = Array.from(el.querySelectorAll('span'));
            if (spans.length > 0) {
                return spans.map(s => s.innerText).join(' ').trim();
            }
        }
        return null;
    }"""

    for tile in tiles:
        try:
            exp_label = tile.evaluate(_exp_js)
            if exp_label is not None:
                label = exp_label
            else:
                gm = tile.locator(".GeneticallyModified").first
                if gm.count() > 0:
                    try:
                        label = gm.evaluate(_walker_js) or ""
                    except Exception:
                        label = gm.inner_text()
                else:
                    label = tile.get_attribute("aria-label") or tile.inner_text() or ""
            if label and label.strip():
                clean_label = " ".join(label.replace("\n", " ").split())
                if clean_label.lower().rstrip(",.!") in ("options", "option"):
                    continue
                options.append(clean_label)
        except Exception:
            continue

    # Drag-and-drop questions: options are draggable tiles in .parking-lot
    if not options:
        if root_locator is not None:
            drag_tiles = root_locator.locator(".parking-lot .draggable-tile").all()
        else:
            drag_tiles = page.locator(
                ".question-and-submission-view .parking-lot .draggable-tile, "
                ".ixl-practice-crate .parking-lot .draggable-tile"
            ).all()
        for tile in drag_tiles:
            try:
                exp_label = tile.evaluate(_exp_js)
                if exp_label is not None:
                    label = exp_label
                else:
                    try:
                        label = tile.evaluate(_walker_js) or ""
                    except Exception:
                        label = tile.inner_text()
                clean_label = " ".join(label.replace("\n", " ").split())
                if clean_label.lower().rstrip(",.!") in ("options", "option"):
                    continue
                options.append(clean_label)
            except Exception:
                continue

    # Sorting drag-and-drop: tiles in .ddItemBankDropSlot
    if not options:
        if root_locator is not None:
            bank_slots = root_locator.locator(".ddItemBankDropSlot").all()
        else:
            bank_slots = page.locator(
                ".question-and-submission-view .ddItemBankDropSlot, "
                ".ixl-practice-crate .ddItemBankDropSlot"
            ).all()
        for slot in bank_slots:
            try:
                exp_label = slot.evaluate(_exp_js)
                if exp_label is not None:
                    label = exp_label
                else:
                    content = slot.locator(".itemContent").first
                    target = content if content.count() > 0 else slot
                    label = target.evaluate(_walker_js) or ""
                clean_label = " ".join(label.replace("\n", " ").split())
                if clean_label.lower().rstrip(",.!") in ("options", "option"):
                    continue
                options.append(clean_label)
            except Exception:
                continue

    # Expression building tiles: .expression-tile-parking-space
    if not options:
        if root_locator is not None:
            exp_tiles = root_locator.locator(".expression-tile-parking-space").all()
        else:
            exp_tiles = page.locator(
                ".question-and-submission-view .expression-tile-parking-space, "
                ".ixl-practice-crate .expression-tile-parking-space"
            ).all()
        if exp_tiles:
            numbers = []
            operators = []
            for tile in exp_tiles:
                try:
                    expr = tile.locator(".expression-tile").first
                    target = expr if expr.count() > 0 else tile
                    
                    label = target.evaluate(_walker_js) or target.inner_text() or ""
                    clean_label = " ".join(label.replace("\n", " ").split())
                    if not clean_label.strip():
                        continue
                    
                    tile_type = tile.get_attribute("data-type") or ""
                    tile_class = tile.get_attribute("class") or ""
                    if "NUMBER" in tile_type.upper() or "number" in tile_class.lower():
                        numbers.append(clean_label)
                    elif "OPERATOR" in tile_type.upper() or "operator" in tile_class.lower():
                        operators.append(clean_label)
                    else:
                        if clean_label in ["+", "-", "=", "x", "×", "÷", "/", "−"]:
                            operators.append(clean_label)
                        else:
                            numbers.append(clean_label)
                except Exception:
                    continue

            seen_nums = set()
            uniq_nums = []
            for n in numbers:
                if n not in seen_nums:
                    seen_nums.add(n)
                    uniq_nums.append(n)
            
            seen_ops = set()
            uniq_ops = []
            for o in operators:
                if o not in seen_ops:
                    seen_ops.add(o)
                    uniq_ops.append(o)
            
            lines = []
            if uniq_nums:
                lines.append(" ".join(uniq_nums))
            if uniq_ops:
                lines.append(" ".join(uniq_ops))
            if lines:
                options.append("\n".join(lines))

    seen, unique = set(), []
    for o in options:
        if o not in seen:
            seen.add(o)
            unique.append(o)
    return "\n".join(unique)


def extract_correct_answer(page):
    try:
        if page.locator(".react-gc-number-button-grid").count() > 0:
            containers = page.locator(".react-gc-number-button-grid .number-button-container").all()
            for container in containers:
                try:
                    title_el = container.locator("svg title").first
                    if title_el.count() > 0:
                        if title_el.text_content().strip().lower() == "selected number button":
                            val = container.inner_text().strip()
                            if val: return val
                except Exception:
                    pass
    except Exception:
        pass

    answer = ""
    box = page.locator(".answer-box .correct-answer").first
    if box.count() == 0:
        box = page.locator(".answer-box").first

    try:
        tiles = box.locator(".SelectableTile").all() if box.count() > 0 else []
        for tile in tiles:
            sr      = tile.locator(".sr-only").first
            sr_text = sr.inner_text().strip() if sr.count() > 0 else ""
            cls     = tile.get_attribute("class") or ""
            if sr_text.lower().startswith("correct answer") or "selected" in cls.split():
                gm = tile.locator(".GeneticallyModified").first
                if gm.count() > 0:
                    try:
                        walker_js = f"""
                        el => {{
                            {_MATH_WALKER_JS}
                            walk(el);
                            return out.trim();
                        }}
                        """
                        val = gm.evaluate(walker_js)
                    except Exception:
                        val = " ".join(gm.inner_text().replace("\n", " ").split())
                    val = " ".join((val or "").replace("\n", " ").split())
                    if val:
                        answer = val
                        break
    except Exception:
        pass

    if not answer and box.count() > 0:
        try:
            inputs = box.locator("input.fillIn").all()
            vals = []
            for inp in inputs:
                v = inp.input_value() or inp.get_attribute("value") or ""
                if v.strip():
                    vals.append(v.strip())
            if vals:
                answer = ", ".join(vals)
        except Exception:
            pass

    if not answer and box.count() > 0:
        try:
            txt    = box.inner_text().replace("\n", " ").strip()
            answer = " ".join(txt.split())
        except Exception:
            pass

    if answer:
        for phrase in ["Option,", "Correct answer,"]:
            answer = answer.replace(phrase, "")
        answer = answer.strip()
    return answer

#  DIAGRAM EXTRACTION

def _safe_bbox(element):
    try:
        # Returns page-relative coordinates and dimensions of web element
        return element.evaluate("""el => {
            const rect = el.getBoundingClientRect();
            return {
                x: rect.left + window.scrollX,
                y: rect.top + window.scrollY,
                width: rect.width,
                height: rect.height
            };
        }""")
    except Exception:
        return None


# def _is_junk_label(label: str) -> bool:
#     return (label or "").strip().lower() in JUNK_LABELS

# Used to ignore microscopic elements that hinder the scraper.
def _is_too_small(bb) -> bool:
    if bb is None:
        return True
    w, h = bb["width"], bb["height"]
    if w < 1 or h < 1:
        return True
    if w <= ICON_MAX and h <= ICON_MAX:
        return True
    return False

# Logic to eliminate duplicate images
def _boxes_overlap(a, b, threshold=0.70):
    if a is None or b is None:
        return False
    ax1, ay1 = a["x"], a["y"]
    ax2, ay2 = ax1 + a["width"],  ay1 + a["height"]
    bx1, by1 = b["x"], b["y"]
    bx2, by2 = bx1 + b["width"],  by1 + b["height"]
    inter_x = max(0, min(ax2, bx2) - max(ax1, bx1))
    inter_y = max(0, min(ay2, by2) - max(ay1, by1))
    inter_area = inter_x * inter_y
    b_area = b["width"] * b["height"]
    if b_area <= 0:
        return False
    return (inter_area / b_area) >= threshold


def _wait_for_element_painted(element, retries=18, delay_ms=50):
    for _ in range(retries):
        bb = _safe_bbox(element)
        if bb and bb["width"] > 1 and bb["height"] > 1:
            return bb
        time.sleep(delay_ms / 1000)
    return None


# Screenshot padding (px) added above and below the element's layout bbox to
# capture visual overflow from negative CSS margins (cube SVGs: -30px top/bottom)
# and absolutely-positioned children above the element top (number-line +1 labels
# sit at top:-23px). Clamped to viewport so it never wraps around.
_SCREENSHOT_VPAD = 32
_SCREENSHOT_HPAD = 4


def _screenshot_element(page, element, path):
    try:
        page.wait_for_timeout(300)
        # Use element.bounding_box() (Playwright built-in) — returns viewport-
        # relative coordinates, which is what page.screenshot(clip=...) expects.
        # This avoids the systematic misalignment that element.screenshot() has
        # when the element's visual content overflows its layout bounding box
        # (e.g. cube SVGs with margin:-30px or number-line labels at top:-23px).
        bb = element.bounding_box()
        if bb is None:
            return False
        vw = page.viewport_size["width"]
        vh = page.viewport_size["height"]
        x = max(0, bb["x"] - _SCREENSHOT_HPAD)
        y = max(0, bb["y"] - _SCREENSHOT_VPAD)
        w = min(vw - x, bb["width"]  + _SCREENSHOT_HPAD * 2)
        h = min(vh - y, bb["height"] + _SCREENSHOT_VPAD * 2)
        if w <= 0 or h <= 0:
            return False
        page.screenshot(path=path, clip={"x": x, "y": y, "width": w, "height": h})
        return True
    except Exception as e:
        print(f"     [!] screenshot failed ({path}): {e}")
        return False

# Count number of diagrams
def _collect_units_from_container(container, scope_label):
    for sel in (
        ".guide-counting-clickable-image-container",
        ".vector-image-wrapper[role='img']",
        ".shape",
        ".pie-chart",
    ):
        units = container.locator(sel).all()
        if units:
            print(f"       [{scope_label}] repeating container → {len(units)} units via '{sel}'")
            return units
    return [container]

# Gate: checks ONLY inside .secHdr and .secContent.
# Each scope part queried independently — no comma-joining with signals.
def _question_has_diagram(page):

    for scope_part in Q_SCOPE_PARTS:
        for signal in DIAGRAM_SIGNALS:
            try:
                # Greater than 0 means, diagram found!
                if page.locator(f"{scope_part} {signal}").count() > 0:
                    return True
            except Exception:
                pass
    return False

# checks signals directly on a single option tile locator
def _tile_has_diagram(tile):

    for signal in DIAGRAM_SIGNALS:
        try:
            if tile.locator(signal).count() > 0:
                return True
        except Exception:
            pass
    return False


def _tile_has_media(tile):
    try:
        return tile.evaluate("""tile => {
            if (tile.querySelector('svg')) return true;
            if (tile.querySelector('canvas')) return true;
            if (tile.querySelector('.simple-item-table')) return true;
            const imgs = tile.querySelectorAll('img');
            for (const img of imgs) {
                const src = img.getAttribute('src') || '';
                if (src && !src.includes('spacer.gif')) return true;
            }
            const allEls = tile.querySelectorAll('*');
            for (const el of allEls) {
                const bg = window.getComputedStyle(el).backgroundImage;
                if (bg && bg !== 'none' && bg.includes('url(')) {
                    return true;
                }
                const styleAttr = el.getAttribute('style') || '';
                if (styleAttr.includes('background') && styleAttr.includes('url(')) {
                    return true;
                }
            }
            return false;
        }""")
    except Exception:
        return False



def extract_diagrams_screenshots(page, question_index, skill_name):
    clean_skill = re.sub(r'[\\/*?:"<>|]', "", skill_name)
    slug = clean_skill.replace(" ", "_")[:40]
    ts = int(time.time())

    q_folder_name = f"{slug}_q{question_index + 1}_qdiag_{ts}"
    opt_folder_name = f"{slug}_q{question_index + 1}_optdiag_{ts}"
    q_folder_path = os.path.join(IMAGE_DIR, q_folder_name)
    opt_folder_path = os.path.join(IMAGE_DIR, opt_folder_name)

    q_folder_id = None
    opt_folder_id = None

    # Find active question container to avoid duplicate phantom elements
    _q_candidates = []
    _seen_q_coords = set()
    for _scope in (".question-and-submission-view", ".ixl-practice-crate"):
        try:
            for _el in page.locator(_scope).all():
                _bb = _safe_bbox(_el)
                if _bb is None or _bb["width"] < 1 or _bb["height"] < 1 or _bb["y"] < 0:
                    continue
                _coord = (round(_bb["x"]), round(_bb["y"]), round(_bb["width"]), round(_bb["height"]))
                if _coord not in _seen_q_coords:
                    _seen_q_coords.add(_coord)
                    _q_candidates.append((_bb["y"], _el))
        except Exception:
            pass

    _active_q = None
    if _q_candidates:
        _q_candidates.sort(key=lambda t: t[0])
        _active_q = _q_candidates[0][1]
    else:
        _active_q = page.locator("body")

    q_paths = []
    opt_paths = []

    if _active_q:
        active_content = _active_q.locator(".secContent")
        if active_content.count() == 0:
            active_content = _active_q
        # Check if active question has diagrams
        has_diagram = False
        for signal in DIAGRAM_SIGNALS:
            try:
                if active_content.locator(signal).count() > 0:
                    has_diagram = True
                    break
            except Exception:
                pass

        if has_diagram:
            os.makedirs(q_folder_path, exist_ok=True)
            q_folder_id = _create_drive_folder(q_folder_name, DRIVE_FOLDER_ID)
            q_paths = _extract_from_scope(
                page=page,
                scope_parts=Q_SCOPE_PARTS,
                root_locator=active_content,
                scope_label="Q",
                prefix=f"{slug}_q{question_index + 1}",
                ts=ts,
                save_dir=q_folder_path,
                drive_folder_id=q_folder_id,
            )

        # Process option tiles inside the active question
        active_tiles = _active_q.locator(".SelectableTile").all()
        for t_idx, tile in enumerate(active_tiles):
            tile_class = tile.get_attribute("class") or ""
            if "TEXT" in tile_class.split():
                continue
            if not (_tile_has_diagram(tile) or _tile_has_media(tile)):
                continue

            target_el = tile
            try:
                if tile.locator(".standalone-cube-train-wrapper .horizontal-cell").count() > 0:
                    target_el = tile.locator(".standalone-cube-train-wrapper .horizontal-cell").first
                elif tile.locator(".horizontal-cell").count() > 0:
                    target_el = tile.locator(".horizontal-cell").first
                elif tile.locator(".vector-image").count() > 0:
                    target_el = tile.locator(".vector-image").first
            except Exception:
                pass

            bb = _wait_for_element_painted(target_el)
            if bb and bb["width"] > 2 and bb["height"] > 2:
                idx = len(opt_paths) + 1
                os.makedirs(opt_folder_path, exist_ok=True)
                if opt_folder_id is None:
                    opt_folder_id = _create_drive_folder(opt_folder_name, DRIVE_FOLDER_ID)
                path = os.path.join(opt_folder_path, f"{slug}_q{question_index + 1}_opt{idx}_{ts}.png")
                if _screenshot_element(page, target_el, path):
                    _upload_file_to_drive(path, opt_folder_id)
                    opt_paths.append(path)
                    print(f"       [Opt{idx}] SAVED tile screenshot: {path}")

    # Sorting drag-and-drop bins: find the live binsContainer (topmost y = active
    # question), then screenshot only its direct .bin children. IXL pre-renders
    # upcoming questions below the viewport, so a flat .binsContainer .bin query
    # returns N×3 results; we must scope to the single live container first.
    _bin_candidates = []
    _seen_bin_coords = set()
    for _scope in (".question-and-submission-view", ".ixl-practice-crate"):
        try:
            for _el in page.locator(f"{_scope} .binsContainer").all():
                _bb = _safe_bbox(_el)
                if _bb is None or _bb["width"] < 1 or _bb["height"] < 1 or _bb["y"] < 0:
                    continue
                _coord = (round(_bb["x"]), round(_bb["y"]),
                          round(_bb["width"]), round(_bb["height"]))
                if _coord not in _seen_bin_coords:
                    _seen_bin_coords.add(_coord)
                    _bin_candidates.append((_bb["y"], _el))
        except Exception:
            pass

    # for bins- first, dropArea, last
    if _bin_candidates:
        _bin_candidates.sort(key=lambda t: t[0])
        _active_container = _bin_candidates[0][1]
        bins_to_screenshot = []
        first_bin = _active_container.locator(".bin.first").first
        if first_bin.count() > 0:
            bins_to_screenshot.append(first_bin)
        bins_to_screenshot.extend(_active_container.locator(".bin:not(.first):not(.last)").all())
        last_bin = _active_container.locator(".bin.last").first
        if last_bin.count() > 0:
            bins_to_screenshot.append(last_bin)

        if not bins_to_screenshot:
            bins_to_screenshot = _active_container.locator(".bin").all()

        for b_idx, bin_el in enumerate(bins_to_screenshot):
            try:
                bb = _wait_for_element_painted(bin_el)
                if bb is None or bb["width"] < 2 or bb["height"] < 2:
                    continue
                os.makedirs(opt_folder_path, exist_ok=True)
                if opt_folder_id is None:
                    opt_folder_id = _create_drive_folder(opt_folder_name, DRIVE_FOLDER_ID)
                path = os.path.join(opt_folder_path,
                                    f"{slug}_q{question_index + 1}_bin{b_idx + 1}_{ts}.png")
                if _screenshot_element(page, bin_el, path):
                    _upload_file_to_drive(path, opt_folder_id)
                    opt_paths.append(path)
                    print(f"       [Bin{b_idx + 1}] SAVED bin screenshot: {path}")
            except Exception as e:
                print(f"     [!] bin screenshot failed: {e}")

    q_url = _drive_subfolder_url(q_folder_id) if q_folder_id else ""
    opt_url = _drive_subfolder_url(opt_folder_id) if opt_folder_id else ""
    return q_url, opt_url

def _extract_from_scope(page, scope_parts, root_locator, scope_label, prefix, ts, save_dir=None, drive_folder_id=None):
    target_dir = save_dir if save_dir is not None else IMAGE_DIR
    paths      = []
    seen_boxes = []

    def already_captured(bb):
        if bb is None:
            return False
        for seen in seen_boxes:
            # Standard overlap check: candidate overlaps 70%+ of a seen box
            if _boxes_overlap(seen, bb):
                return True
            # Containment check: candidate is fully inside a seen box (child element).
            # Prevents sub-elements (e.g. individual SVG shapes inside a tenFrames table
            # that was already shot as a whole) from being re-captured.
            if (seen["x"] <= bb["x"] and seen["y"] <= bb["y"]
                    and seen["x"] + seen["width"]  >= bb["x"] + bb["width"]
                    and seen["y"] + seen["height"] >= bb["y"] + bb["height"]):
                return True
        return False

    def do_screenshot(element, tag, layer="?", signal="?"):
        try:
            el_tag = element.evaluate("el => el.tagName.toLowerCase()")
        except Exception:
            el_tag = ""

        try:
            # SVG <g> elements have no valid HTML bounding box — getBoundingClientRect()
            # on a <g> returns coordinates in the SVG's internal coordinate system, not
            # the HTML page. Walk up to the nearest <svg> ancestor which does have a
            # proper HTML bbox and can be screenshotted correctly.
            if el_tag == "g":
                parent_svg = element.locator("xpath=ancestor::svg[1]")
                if parent_svg.count() > 0:
                    element = parent_svg.first
            elif element.locator(".standalone-cube-train-wrapper .horizontal-cell").count() > 0:
                element = element.locator(".standalone-cube-train-wrapper .horizontal-cell").first
            elif "standalone-cube-train-wrapper" in (element.get_attribute("class") or ""):
                if element.locator(".horizontal-cell").count() > 0:
                    element = element.locator(".horizontal-cell").first
            elif "static-cube-train" in (element.get_attribute("class") or ""):
                # Screenshot the static-cube-train div directly — do NOT drill into children
                # because the cube SVGs have negative margins that overflow the div bounds.
                pass
            elif element.locator(".horizontal-cell").count() > 0:
                element = element.locator(".horizontal-cell").first
            elif element.locator(".vector-image").count() > 0:
                element = element.locator(".vector-image").first
        except Exception:
            pass

        bb = _wait_for_element_painted(element)
        if _is_too_small(bb):
            print(f"       [{scope_label}] DROP-small  layer={layer} sig={signal} bb={bb}")
            return
        if already_captured(bb):
            print(f"       [{scope_label}] DROP-overlap layer={layer} sig={signal} "
                  f"bb=({round(bb['x'])},{round(bb['y'])},{round(bb['width'])},{round(bb['height'])})")
            return
        try:
            # This relurns an inline JS object matching elements's actual HTML tag names, classes etc
            # Used to print in console about the image info
            meta = element.evaluate(
                "el => ({tag: el.tagName, cls: el.getAttribute('class'), "
                "role: el.getAttribute('role'), al: el.getAttribute('aria-label')})"
            )
        except Exception as e:
            meta = {"tag": "?", "cls": f"<eval failed: {e}>", "role": "?", "al": "?"}
        idx  = len(paths) + 1
        path = os.path.join(target_dir, f"{prefix}_{tag}_{ts}_{idx}.png")
        if _screenshot_element(page, element, path):
            paths.append(path)
            seen_boxes.append(bb)
            if drive_folder_id:
                _upload_file_to_drive(path, drive_folder_id)
            print(f"       [{scope_label}] SAVED#{idx} layer={layer} sig={signal} "
                  f"tag={meta.get('tag')} cls={meta.get('cls')} role={meta.get('role')} "
                  f"al={meta.get('al')} "
                  f"bb=({round(bb['x'])},{round(bb['y'])},{round(bb['width'])},{round(bb['height'])})")

    def _is_really_visible(el):
        try:
            if not el.is_visible():
                return False
        except Exception:
            return False
        bb = _safe_bbox(el)
        if bb is None:
            return False
        # Coordinates are not (0,0), even more negative
        if bb["y"] < 0 or bb["x"] < 0:
            return False
        if bb["width"] < 1 or bb["height"] < 1:
            return False
        return True

    def _is_inside_option(el):
        if scope_label != "Q":
            return False
        try:
            return el.evaluate("el => el.closest('.SelectableTile, .TileMultipleChoices, .parking-lot, .draggable-tile, .ddItemBankDropSlot, .expression-tile-parking-space, .answer-box') !== null")
        except Exception:
            return False

    def _is_inside_venn_diagram(el):
        """Return True if el is a child/descendant of a venn diagram container (not the container itself).
        Prevents individual shapes inside dragAndDropVennDiagramContainer from getting their own screenshot."""
        try:
            return el.evaluate("""el => {
                if (el.matches('[class*="dragAndDropVennDiagramContainer"]')) return false;
                return el.closest('[class*="dragAndDropVennDiagramContainer"]') !== null;
            }""")
        except Exception:
            return False

    def get_elements(signal):
        if root_locator is not None:
            try:
                return [el for el in root_locator.locator(signal).all()
                        if _is_really_visible(el) and not _is_inside_option(el) and not _is_inside_venn_diagram(el)]
            except Exception:
                return []
        else:
            results     = []
            seen_coords = set()
            for part in scope_parts:
                try:
                    els = page.locator(f"{part} {signal}").all()
                    for el in els:
                        if not _is_really_visible(el) or _is_inside_option(el) or _is_inside_venn_diagram(el):
                            continue
                        bb = _safe_bbox(el)
                        coord_key = (round(bb["x"]), round(bb["y"]),
                                     round(bb["width"]), round(bb["height"]))
                        if coord_key not in seen_coords:
                            seen_coords.add(coord_key)
                            results.append(el)
                except Exception:
                    pass
            return results

    L1_INTEGRATED = [
        # Series-of-components MUST come before vector-image-wrapper so the whole
        # row (both image groups side-by-side) is captured as one unit and each
        # individual vector-image-wrapper inside gets dropped by the containment check.
        "[class*='series-of-components']",
        ".horizontal-scroll-hoc-wrapper",
        ".horizontal-scroll-element-wrapper",
        ".multiplication-model-container",
        ".fractionTopBlockDiv",
        ".open-number-line",
        ".dc-fraction-strip-model",
        "div.table:has([data-testid='area-model-cell'])",
        "table.old-table",
        "table.qTabularGrid",
        "svg:has(g.grid-region)",
        ".gc-cut-shapes",
        "canvas",
        '[role="figure"]',
        ".shape",
        ".vector-image-wrapper",
        ".standalone-cube-train-wrapper",
        ".hundredTable",
        "table:has(img[src*='~media'])",
        # New integrated signals
        ".train-and-item-group",
        ".train-and-element-group",
        ".measurementRegion",
        ".calendar-container",
        ".diagramLabelContainer",
        ".QMMeasurable",
        "[class*='story-book']",
        "[class*='static-cube-train']",
        "[class*='qTable']",
        "[class*='pvmContainer']", 
        "[class*='clockContainer']",
        "[class*='currencyCoinDiv']",
        "[class*='horizontal-scroll']",
        "[class*='dragAndDropContainer']",
        "[class*='SelectableTime']",
        ".simple-item-table"
    ]
    # These can have multiple real instances per question (e.g. two number lines for
    # equivalence questions, or a standalone pie chart) — no phantom de-dupe applied.
    L1_MULTI      = [".graphingBaseContainer", ".pie-chart", ".qPVTable", ".guide-counting-clickable-image-container", "[class*='tenFrames']"]
    L1_REPEATING  = [".guide-counting-qm"]

    for container_sel in L1_INTEGRATED + L1_MULTI + L1_REPEATING:
        containers = get_elements(container_sel)

        # For integrated single-figure diagrams in the QUESTION scope, IXL pre-renders upcoming-question copies stacked below the live one.
        # Keep only the topmost (smallest y) — that's the active diagram.
        # L1_MULTI types are exempt: a question may show several legitimately.
        if container_sel in L1_INTEGRATED and root_locator is None and len(containers) > 1:
            containers = sorted(
                containers,
                key=lambda c: (_safe_bbox(c) or {"y": 1e9})["y"]
            )[:1]
            print(f"       [{scope_label}] {container_sel}: kept topmost of "
                  f"{container_sel} (phantom de-dupe)")

        for container in containers:
            bb_container = _safe_bbox(container)
            if already_captured(bb_container):
                continue
            if container_sel in L1_INTEGRATED or container_sel in L1_MULTI:
                do_screenshot(container, "fig", layer="L1-int", signal=container_sel)
            else:
                units = _collect_units_from_container(container, scope_label)
                for unit in units:
                    do_screenshot(unit, "fig", layer="L1-rep", signal=container_sel)

    return paths

# After submission, screenshot regular diagrams in the answer box and each bin (with placed tiles) as the correct answer.
def _screenshot_answer_diagrams(page, question_index, skill_name, ts):
    clean_skill = re.sub(r'[\\/*?:"<>|]', "", skill_name)
    slug = clean_skill.replace(" ", "_")[:40]
    ans_folder_name = f"{slug}_q{question_index + 1}_ansdiag_{ts}"
    ans_folder_path = os.path.join(IMAGE_DIR, ans_folder_name)
    ans_folder_id = None
    paths = []

    # 1. Extract regular diagrams inside the answer box
    try:
        answer_box = page.locator(".answer-box").first
        if answer_box.count() > 0 and answer_box.is_visible():
            tiles = answer_box.locator(".SelectableTile")
            is_tile = False
            if tiles.count() > 0:
                selected_tile = answer_box.locator(".SelectableTile.selected")
                if selected_tile.count() > 0:
                    answer_box = selected_tile.first
                    is_tile = True
                else:
                    answer_box = tiles.first
                    is_tile = True

            if is_tile:
                tile_class = answer_box.get_attribute("class") or ""
                if "TEXT" not in tile_class.split() and (_tile_has_diagram(answer_box) or _tile_has_media(answer_box)):
                    target_el = answer_box
                    try:
                        if answer_box.locator(".standalone-cube-train-wrapper .horizontal-cell").count() > 0:
                            target_el = answer_box.locator(".standalone-cube-train-wrapper .horizontal-cell").first
                        elif answer_box.locator(".horizontal-cell").count() > 0:
                            target_el = answer_box.locator(".horizontal-cell").first
                        elif answer_box.locator(".vector-image").count() > 0:
                            target_el = answer_box.locator(".vector-image").first
                    except Exception:
                        pass
                    
                    bb = _wait_for_element_painted(target_el)
                    if bb and bb["width"] > 2 and bb["height"] > 2:
                        os.makedirs(ans_folder_path, exist_ok=True)
                        ans_folder_id = _create_drive_folder(ans_folder_name, DRIVE_FOLDER_ID)
                        path = os.path.join(ans_folder_path, f"{slug}_q{question_index + 1}_ans_tile_{ts}.png")
                        if _screenshot_element(page, target_el, path):
                            _upload_file_to_drive(path, ans_folder_id)
                            paths.append(path)
                            print(f"       [Ans] SAVED answer tile screenshot: {path}")
            else:
                has_diagram = False
                for signal in DIAGRAM_SIGNALS:
                    try:
                        if answer_box.locator(signal).count() > 0:
                            has_diagram = True
                            break
                    except Exception:
                        pass

                if has_diagram:
                    os.makedirs(ans_folder_path, exist_ok=True)
                    ans_folder_id = _create_drive_folder(ans_folder_name, DRIVE_FOLDER_ID)
                    ans_diag_paths = _extract_from_scope(
                        page=page,
                        scope_parts=[],
                        root_locator=answer_box,
                        scope_label="Ans",
                        prefix=f"{slug}_q{question_index + 1}_ans",
                        ts=ts,
                        save_dir=ans_folder_path,
                        drive_folder_id=ans_folder_id,
                    )
                    paths.extend(ans_diag_paths)
    except Exception as e:
        print(f"     [!] answer diagram screenshot failed: {e}")

    # 2. Extract answer bins
    try:
        all_candidates = []
        seen_coords = set()
        for scope in (".answer-box", ".question-and-submission-view", ".ixl-practice-crate"):
            try:
                for el in page.locator(f"{scope} .binsContainer").all():
                    bb = _safe_bbox(el)
                    if bb is None or bb["width"] < 1 or bb["height"] < 1 or bb["y"] < 0:
                        continue
                    coord = (round(bb["x"]), round(bb["y"]),
                             round(bb["width"]), round(bb["height"]))
                    if coord not in seen_coords:
                        seen_coords.add(coord)
                        all_candidates.append((bb["y"], el))
            except Exception:
                pass

        if all_candidates:
            with_tiles = []
            for y, el in all_candidates:
                try:
                    if el.locator(".bin .draggableElement").count() > 0:
                        with_tiles.append((y, el))
                except Exception:
                    pass

            candidates = with_tiles if with_tiles else all_candidates
            candidates.sort(key=lambda t: t[0])
            # selects the topmost container to take screenshot of valid image and ignore duplicates
            container = candidates[0][1]

            bins_to_screenshot = []
            first_bin = container.locator(".bin.first").first
            if first_bin.count() > 0:
                bins_to_screenshot.append(first_bin)
            bins_to_screenshot.extend(container.locator(".bin:not(.first):not(.last)").all())
            last_bin = container.locator(".bin.last").first
            if last_bin.count() > 0:
                bins_to_screenshot.append(last_bin)

            if not bins_to_screenshot:
                bins_to_screenshot = container.locator(".bin").all()

            for b_idx, bin_el in enumerate(bins_to_screenshot):
                bb = _wait_for_element_painted(bin_el)
                if bb is None or bb["width"] < 2 or bb["height"] < 2:
                    continue
                os.makedirs(ans_folder_path, exist_ok=True)
                if ans_folder_id is None:
                    ans_folder_id = _create_drive_folder(ans_folder_name, DRIVE_FOLDER_ID)
                path = os.path.join(ans_folder_path,
                                    f"{slug}_q{question_index + 1}_ans_bin{b_idx + 1}_{ts}.png")
                if _screenshot_element(page, bin_el, path):
                    _upload_file_to_drive(path, ans_folder_id)
                    paths.append(path)
                    print(f"       [AnsBin{b_idx + 1}] SAVED answer bin: {path}")
    except Exception as e:
        print(f"     [!] answer bin screenshot failed: {e}")

    return _drive_subfolder_url(ans_folder_id) if ans_folder_id else ""


def extract_and_advance(page, category_name, skill_name, serial_tracker):
    previous_question_text = ""

    for i in range(QUESTIONS_PER_SKILL):
        print(f"  -> Processing Question {i+1}...")

        try:
            page.wait_for_selector(
                ".ixl-practice-crate, .math.section, .question-component",
                state="visible", timeout=15000
            )
            page.wait_for_timeout(800)

            for _ in range(41):
                candidate = extract_question_text(page)
                if candidate and candidate != previous_question_text:
                    break
                page.wait_for_timeout(500)
        except Exception as e:
            print(f"     [!] Failed to stabilize DOM for question {i+1}: {e}")
            break

        question_text = extract_question_text(page)
        options_text  = extract_options(page)
        q_diagrams, opt_diagrams = extract_diagrams_screenshots(page, i, skill_name)
        previous_question_text = question_text

        if not question_text:
            print(f"     [!] WARNING: empty question text on Q{i+1}")


        correct_answer = ""
        try:
            # First Submit button - for incomplete answers
            submit_btn = page.locator(
                'button[data-cy="question-submit-button"], div.question button.submit'
            ).first
            submit_btn.wait_for(state="visible", timeout=10000)
            submit_btn.click()

            # Pop-up Submit button - for incomplete answers
            popup_btn = page.locator(
                'button[data-cy="incomplete-answer-popover-submit-button"]'
            ).first
            popup_btn.wait_for(state="visible", timeout=5000)
            popup_btn.click()

            page.wait_for_selector(".answer-box", state="visible", timeout=10000)
            page.wait_for_timeout(600)
            correct_answer = extract_correct_answer(page)
        except Exception as e:
            print(f"     [!] Could not read correct answer on Q{i+1}: {e}")

        ts_ans = int(time.time())
        ans_diagrams = _screenshot_answer_diagrams(page, i, skill_name, ts_ans)

        # row_data = columns A-E, G, I (7 values)
        row_vals = [
            serial_tracker[0] if i == 0 else "",
            category_name     if i == 0 else "",
            skill_name        if i == 0 else "",
            i + 1,
            question_text,
            options_text,
            correct_answer
        ]
        append_to_excel(row_vals, q_diagrams, opt_diagrams, ans_diagram_url=ans_diagrams)

        if i < QUESTIONS_PER_SKILL - 1:
            try:
                # Got it button
                got_it_btn = page.locator(
                    'button:has-text("Got it"), .next-problem button'
                ).first
                got_it_btn.wait_for(state="visible", timeout=15000)
                got_it_btn.click()
                got_it_btn.wait_for(state="hidden", timeout=10000)
            except Exception as e:
                print(f"     [!] Transition error on Q{i+1}: {e}")
                break
        else:
            try:
                got_it_btn = page.locator(
                    'button:has-text("Got it"), .next-problem button'
                ).first
                if got_it_btn.count() > 0 and got_it_btn.is_visible():
                    got_it_btn.click()
                    page.wait_for_timeout(500)
            except Exception:
                pass

    serial_tracker[0] += 1

def _choose_mode():
    print("\n  IXL Scraper\n")
    print("  [1]  Scrape entire website from scratch")
    print(f"  [2]  Start from a specific skill URL - {' '.join(START_URL.split('/')[-1].split('.')[0].split('-')).title()}")
    while True:
        choice = input("\n  Enter mode (1 or 2): ").strip()
        if choice in ("1", "2"):
            return int(choice)
        print("  Invalid choice. Please enter 1 or 2.")


def run_scraper():
    mode = _choose_mode()

    if mode == 2:
        # Normalise START_URL to just its path for comparison
        start_path = urllib.parse.urlparse(START_URL).path.rstrip("/")
        print(f"\n[Mode 2] Will skip skills until: {START_URL}")
        reached_start = False
    else:
        start_path    = None
        reached_start = True  # Mode 1: start immediately

    _init_drive_service()
    setup_dir()
    init_excel()

    with sync_playwright() as p:
        print("\nInitializing browser context...")
        browser = p.chromium.launch(headless=False)
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                       "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
        )
        page = context.new_page()

        print("Authenticating...")
        page.goto(LOGIN_URL)

        # Fill in credentials
        page.fill("input[type='email'], input[name='username']", EMAIL)
        page.fill("input[type='password'], input[name='password']", PASSWORD)
        page.keyboard.press("Enter")
        page.wait_for_timeout(6000)

        print(f"Navigating to target directory: {TARGET_URL}")
        page.goto(TARGET_URL)
        page.wait_for_selector("div.skill-tree-category", state="attached", timeout=15000)

        total_categories = page.locator("div.skill-tree-category").count()
        print(f"Found {total_categories} category blocks.")

        serial_tracker = [1]

        for cat_index in range(total_categories):
            # Re-fetch count in case DOM changed after navigation
            current_count = page.locator("div.skill-tree-category").count()
            
            scroll_attempts = 0
            while cat_index >= current_count and scroll_attempts < 10:
                print(f"  [Scroll] Category index {cat_index} not in DOM (current count: {current_count}). Scrolling to load...")
                try:
                    page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                    page.wait_for_timeout(1000)
                except Exception:
                    pass
                current_count = page.locator("div.skill-tree-category").count()
                scroll_attempts += 1

            if cat_index >= current_count:
                print(f"  [!] Category index {cat_index} out of range (only {current_count} found). Stopping.")
                break

            cat_block = page.locator("div.skill-tree-category").nth(cat_index)

            # Scroll the category block into view to trigger lazy rendering
            try:
                cat_block.scroll_into_view_if_needed()
                page.wait_for_timeout(300)
            except Exception:
                pass

            header_loc = cat_block.locator(".skill-tree-skills-header").first
            if not header_loc.is_visible():
                # Try waiting briefly for visibility after scroll
                try:
                    header_loc.wait_for(state="visible", timeout=3000)
                except Exception:
                    print(f"  [!] Category {cat_index} header not visible after scroll — skipping.")
                    continue

            code_text = (header_loc.locator(".category-code").inner_text().strip()
                         if header_loc.locator(".category-code").count() > 0 else "")
            name_text = (header_loc.locator(".category-name").inner_text().strip()
                         if header_loc.locator(".category-name").count() > 0 else "")
            full_category_name = f"{code_text} {name_text}".strip()

            skill_links = cat_block.locator("a.skill-tree-skill-link")
            skill_count = skill_links.count()

            for skill_index in range(skill_count):
                current_skill = (page.locator("div.skill-tree-category").nth(cat_index)
                                 .locator("a.skill-tree-skill-link").nth(skill_index))

                title_span = current_skill.locator("span.skill-tree-skill-name").first
                skill_name = (title_span.inner_text().strip()
                              if title_span.count() > 0
                              else current_skill.inner_text().strip())

                # Mode 2: skip skills until we hit the START_URL
                if not reached_start:
                    skill_href = (current_skill.get_attribute("href") or "").rstrip("/")
                    if skill_href == start_path:
                        reached_start = True
                        print(f"\n[Mode 2] Resuming from: {skill_name}")
                    else:
                        print(f"  [SKIP] {full_category_name} / {skill_name}")
                        continue

                print(f"\n[{full_category_name}] Entering skill: {skill_name}")

                current_skill.scroll_into_view_if_needed()
                current_skill.click()

                # This runs 3 times
                extract_and_advance(page, full_category_name, skill_name, serial_tracker)

                print("  -> Retreating to main directory...")
                try:
                    breadcrumb = page.locator(
                        'a.breadcrumb-link:has-text("Third grade"), '
                        'a.breadcrumb-link[href="/math/grade-3"], '
                        'a.breadcrumb-link[href="/maths/class-iii"]'
                    ).first
                    breadcrumb.wait_for(state="visible", timeout=10000)
                    breadcrumb.click()
                    page.wait_for_selector("div.skill-tree-category",
                                           state="attached", timeout=15000)
                except Exception as e:
                    print(f"  [!] Failed to use breadcrumb. Forcing URL reload. {e}")
                    page.goto(TARGET_URL)
                    page.wait_for_selector("div.skill-tree-category",
                                           state="attached", timeout=15000)

        print("\nAll topics processed. Closing driver.")
        browser.close()

if __name__ == "__main__":
    run_scraper()
