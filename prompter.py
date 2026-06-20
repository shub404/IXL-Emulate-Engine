from openpyxl import load_workbook
from groq import AsyncGroq
import asyncio
import csv
import os
import re

API_KEY     = "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"
INPUT_FILE  = "ixl_grade3_questions.xlsx"
OUTPUT_FILE = "ixl_grade3_questions_updated.csv"
BATCH_SIZE  = 5   # rows per API call
MAX_CONCURRENT = 3  # parallel batches at once

wb = load_workbook(INPUT_FILE)
ws = wb.active

headers    = {cell.value: cell.column for cell in ws[1]}
col_q      = headers["Question Text"]
col_o      = headers["Question Options"]
col_a      = headers["Correct Answer"]
header_row = [cell.value for cell in ws[1]]

# Resume: count rows already in CSV
start_data_row = 2
if os.path.exists(OUTPUT_FILE):
    with open(OUTPUT_FILE, 'r', newline='', encoding='utf-8') as f:
        rows_written = sum(1 for _ in csv.reader(f)) - 1
    if rows_written > 0:
        start_data_row = 2 + rows_written
        print(f"Resuming from row {start_data_row} (skipping {rows_written} already written rows)")

PROMPT_TEMPLATE = """\
You are a warm, patient Grade 3 math teacher reading questions aloud to 8-year-old students in a classroom.
This output will be used to train a Text-to-Speech model, so it must sound exactly like natural spoken English — never like written text.

RULE 1 — Pauses and rhythm (for TTS prosody):
- Use commas for short pauses and ellipses (...) for longer, deliberate pauses.
- Pause before every number, math term, or key comparison so students have time to process: "There are... four equal parts."
- Pause between the problem setup and the final question being asked.
- Slow down on shape names, quantities, and operations — treat them as important beats.

RULE 2 — Short sentences with complete information:
- Every sentence must be 15 words or fewer. Break one long idea into 2 or 3 short sentences.
- Every number, shape, quantity, unit, condition, and math detail from the raw question MUST appear in your output — nothing can be dropped, summarized, or implied. If the raw question mentions 4 sides, say 4 sides. If it mentions a specific unit, say that unit.
- Spread details across multiple short sentences rather than packing them into one.

RULE 3 — Options format:
- Format the options cleanly as "Option A: [text]", "Option B: [text]", etc.

RULE 4 — Answer announcement:
- Reveal the answer with a warm lead-in, a pause, then the answer, then brief encouragement.
- Pattern: "[Warm phrase]! The correct answer is... [answer]. [Short encouragement]."
- The warm phrase must vary — use "Wonderful!", "That's right!", "Let's see...", "And the answer is" etc.
- The encouragement must vary — use "Great thinking!", "Well done!", "You're doing so well!", "Excellent work!" etc.
- Never write "Correct Answer:" as a label — it must sound like a teacher speaking.

RULE 5 — No non-speech symbols:
- Never use: slashes, asterisks, brackets, hyphens as bullets, or markdown. Only words and standard punctuation that sound natural when spoken aloud.

Convert each of the {n} rows below. Output exactly {n} results.
Separate each result with ||| on its own line.
Within each result, separate the 3 parts with |~|
Do not include any extra text, numbering, or labels — only the results.

{rows_text}"""


def build_batch_prompt(batch_items):
    rows_text = []
    for i, (_, q, o, a, _) in enumerate(batch_items, 1):
        rows_text.append(f"--- ROW {i} ---\nQuestion: {q}\nOptions: {o}\nAnswer: {a}")
    return PROMPT_TEMPLATE.format(n=len(batch_items), rows_text="\n\n".join(rows_text))


async def process_batch(client, batch_items, semaphore):
    """Returns list of processed row_data in same order as batch_items."""
    results = [list(rd) for _, _, _, _, rd in batch_items]

    non_empty = [(i, item) for i, item in enumerate(batch_items) if item[1]]
    if not non_empty:
        return results

    async with semaphore:
        prompt = build_batch_prompt([item for _, item in non_empty])

        while True:
            try:
                response = await client.chat.completions.create(
                    model="llama-3.3-70b-versatile",
                    messages=[{"role": "user", "content": prompt}],
                    temperature=0.2,
                    max_tokens=600 * len(non_empty),
                )
                result_text = response.choices[0].message.content.strip()
                raw = [r.strip() for r in result_text.split('|||') if r.strip()]

                for i, (orig_idx, (row_idx, _, _, _, row_data)) in enumerate(non_empty):
                    if i < len(raw):
                        parts = raw[i].split('|~|')
                        if len(parts) == 3:
                            rd = list(row_data)
                            rd[col_q - 1] = parts[0].strip()
                            rd[col_o - 1] = parts[1].strip()
                            rd[col_a - 1] = parts[2].strip()
                            results[orig_idx] = rd
                            print(f"  Row {row_idx - 1} done")
                        else:
                            print(f"  Row {row_idx - 1} bad format — keeping original")
                    else:
                        print(f"  Row {row_idx - 1} missing from response — keeping original")

                return results

            except Exception as e:
                err_str = str(e)
                if "429" in err_str or "rate limit" in err_str.lower():
                    if "day" in err_str.lower():
                        print("\n[DAILY QUOTA EXHAUSTED] Resets at midnight UTC. Returning originals for this batch.")
                        return results
                    wait = 65
                    match = re.search(r"try again in (\d+\.?\d*)s", err_str.lower())
                    if match:
                        wait = int(float(match.group(1))) + 5
                    print(f"  Rate limit — waiting {wait}s before retry...")
                    await asyncio.sleep(wait)
                else:
                    print(f"  Error on batch: {e} — keeping originals")
                    return results


async def main():
    client = AsyncGroq(api_key=API_KEY)
    semaphore = asyncio.Semaphore(MAX_CONCURRENT)

    # Collect all rows to process
    all_rows = []
    for row_idx in range(start_data_row, ws.max_row + 1):
        q  = str(ws.cell(row=row_idx, column=col_q).value or "").strip()
        o  = str(ws.cell(row=row_idx, column=col_o).value or "").strip()
        a  = str(ws.cell(row=row_idx, column=col_a).value or "").strip()
        rd = [ws.cell(row=row_idx, column=col).value for col in range(1, len(header_row) + 1)]
        all_rows.append((row_idx, q, o, a, rd))

    batches = [all_rows[i:i+BATCH_SIZE] for i in range(0, len(all_rows), BATCH_SIZE)]
    print(f"Rows to process: {len(all_rows)} | Batches: {len(batches)} | {MAX_CONCURRENT} concurrent x {BATCH_SIZE} rows each\n")

    csv_mode = 'a' if start_data_row > 2 else 'w'
    csv_file = open(OUTPUT_FILE, csv_mode, newline='', encoding='utf-8')
    writer   = csv.writer(csv_file)
    if csv_mode == 'w':
        writer.writerow(header_row)

    updates       = 0
    total_written = 0

    # Process in chunks of MAX_CONCURRENT batches, write after each chunk
    for chunk_start in range(0, len(batches), MAX_CONCURRENT):
        chunk = batches[chunk_start:chunk_start + MAX_CONCURRENT]
        tasks = [process_batch(client, batch, semaphore) for batch in chunk]
        chunk_results = await asyncio.gather(*tasks)

        for batch_items, processed_rows in zip(chunk, chunk_results):
            for (row_idx, q, o, a, orig_rd), proc_rd in zip(batch_items, processed_rows):
                writer.writerow(proc_rd)
                total_written += 1
                if q and proc_rd[col_q - 1] != orig_rd[col_q - 1]:
                    updates += 1

        csv_file.flush()
        print(f"[Saved] {total_written}/{len(all_rows)} rows written — {updates} updated by LLM\n")

    csv_file.close()
    print(f"Done. {updates} rows updated. Saved to {OUTPUT_FILE}")


if __name__ == "__main__":
    asyncio.run(main())
